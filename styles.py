#Styling functions in Excel

from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))


def fillColor(min_r, max_r, min_c, max_c, colour, ws):

    for rows1 in ws.iter_rows(min_row=min_r, max_row=max_r, min_col= min_c, max_col = max_c):
        for cell in rows1:
            cell.fill = PatternFill(start_color = colour, end_color = colour, fill_type = "solid")
            cell.font = Font(name= 'Helvetica', size=12)

def applyFont(r, min_c, max_c, ws):

    for rows1 in ws.iter_rows(min_row=r, max_row=r, min_col= min_c, max_col = max_c):
        for cell in rows1:
            cell.font = Font(name= 'Helvetica', size=12)

def border(min_r, max_r, min_c, max_c, ws):

    for rows1 in ws.iter_rows(min_row=min_r, max_row=max_r, min_col= min_c, max_col = max_c):
        for cell in rows1:
            cell.border = thin_border


def align(ws):

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text = True )

    for col in ws.columns:
        max_length = 0
        column = col[0].column  
        for cell in col:
            try:  
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.3
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width