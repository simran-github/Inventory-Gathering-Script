import apiCall, styles
from openpyxl.styles import Font


#Generate PC Overview Excel Sheet of all Clusters 
def getHostDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating Host Details:")

    ws = wb.create_sheet("Host Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:Y{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'Host Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+15, '7855FA', ws)
    r += 1

    global hostsIPs
    hostsIPs = {}
    for clus in clustersData['entities']:
        
        if clus['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = clus['status']['name']
        print(f"Generating Host Details for {clus_name}")
        ws.merge_cells(f'A{r}:Y{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} Hosts'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+15, 'A9D0F5', ws)
        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = clus['metadata']['uuid']
    

        clus_ip = clus['status']['resources']['network']['external_ip']
       

        url = f"https://{clus_ip}:9440/api/lcm/v4.0.a1/resources/entities"

        lcmres = apiCall.apiV4(url, peAuthKey)

        ws.append(["Node Name","Node Serial Number", "Node Model","Block SN", "CPU Model","BIOS","BMC", "Number of Cores", "Number of Sockets", "Memory Capacity (GiB)", "Memory Usage (%)", "CPU Capacity (GHz)", "CPU Usage (%)", "Storage Capacity (TiB)", "Storage Usage (GiB)", "Disk IOPS", "Disk IO B/W", "Disk IO Latency", "Node UUID","Hypervisor IP", "CVM IP", "IPMI IP", "Location", "Cluster Name", "Contract Status"])
        styles.fillColor(r, r, c, c+24, 'D3D3D3', ws)
        styles.applyFont(r, c, c+24, ws)
        r += 1

        url = f"https://{pc_name}:9440/api/nutanix/v3/hosts/list"
      
        dataHost = apiCall.postApiCall(url, pcAuthKey)

        #Loop through each host

        for entity in dataHost['entities']:
                                         
              if entity['status']['cluster_reference']['uuid'] == clus_uuid:

                    node_uuid = entity['metadata']['uuid']

                    url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v2.0/hosts/{node_uuid}"
                    host2Data = apiCall.getApiCall(url, peAuthKey, clus_uuid)

                    node_name = '-'
                    if 'name' in entity['status']:
                      node_name = entity['status']['name']

                    if host2Data['host_in_maintenance_mode']:

                      cell = ws.cell(row=r, column=1)
                      cell.value = f"{node_name}"

                      ws.merge_cells(f'B{r}:Y{r}')
                      cell = ws.cell(row=r, column=2)
                      cell.value = "Host Under Maintenance Mode"
                      r += 1
                      hyp_ip = entity['status']['resources']['hypervisor']['ip']
                      hostsIPs[hyp_ip] = node_name
                      continue

                  
                    block_no = entity['status']['resources']['block']['block_serial_number']
                    


                    node_sno = entity['status']['resources']['serial_number']

                    node_model = entity['status']['resources']['block']['block_model']

                    cpu_model = '-'
                    if 'cpu_model' in entity['status']['resources']:
                      cpu_model = entity['status']['resources']['cpu_model']

                    

                    bios = ""
                    bmc = ""

                    for k in lcmres['entities']:

                      if k['entityClass'] == 'BIOS' and k['locationId'] == f'node:{node_uuid}':
                        bios = f"Model: {k['entityModel']}, Version: {k['version']}"


                      if k['entityClass'] == 'BMCs' and k['locationId'] == f'node:{node_uuid}':
                        bmc = f"Model: {k['entityModel']}, Version: {k['version']}"

                    if bios == "" and bmc == "":

                      bios = "Third Party Hardware. Please fill info manually"
                      bmc = "Third Party Hardware. Please fill info manually" 

                    cores = '-'
                    if 'num_cpu_cores' in entity['status']['resources']:
                      cores = entity['status']['resources']['num_cpu_cores']

                    sockets = '-'
                    if 'num_cpu_sockets' in entity['status']['resources']:
                      sockets = entity['status']['resources']['num_cpu_sockets']

                    memory = '-'
                    if 'memory_capacity_mib' in entity['status']['resources']:
                      memory = round(entity['status']['resources']['memory_capacity_mib'] / 1024, 2)

                    

                    
                    cpu = round(int(host2Data["cpu_capacity_in_hz"]) / 1000000000, 2)
                    cpu_usage = round(int(host2Data["stats"]["hypervisor_cpu_usage_ppm"]) / 10000, 2)

                    mem_usage = round(int(host2Data["stats"]["hypervisor_memory_usage_ppm"]) / 10000, 2)

                    storage = round(int(host2Data["usage_stats"]["storage.capacity_bytes"]) / 1099511627776, 2)

                    storage_usage = round(int(host2Data["usage_stats"]["storage.usage_bytes"]) / 1073741824, 2)

                    disk_iops = int(host2Data["stats"]["num_iops"])

                    disk_bw = int(host2Data["stats"]["io_bandwidth_kBps"])

                    disk_latency = round(int(host2Data["stats"]["avg_io_latency_usecs"]) / 1000, 2)



                    hyp_ip = entity['status']['resources']['hypervisor']['ip']

                    if entity['status']['resources']['host_type'] == 'COMPUTE_ONLY':
                      cvm_ip = "NA. Compute-only Node"

                    else:

                      cvm_ip = entity['status']['resources']['controller_vm']['ip'] 

                    ipmi_ip = entity['status']['resources']['ipmi']['ip']


                    ws.append([node_name, node_sno, node_model, block_no, cpu_model, bios, bmc, cores, sockets, memory, mem_usage, cpu, cpu_usage, storage, storage_usage, disk_iops, disk_bw, disk_latency, node_uuid, hyp_ip, cvm_ip, ipmi_ip, "", clus_name, ""])
                    styles.applyFont(r, c, c+24, ws)
                    r += 1
                    hostsIPs[hyp_ip] = node_name

    styles.align(ws)
    styles.border(1, r-1, 1, 25, ws)