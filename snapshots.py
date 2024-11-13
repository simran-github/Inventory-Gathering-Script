import apiCall, styles
from datetime import datetime, timedelta
from openpyxl.styles import Font


#Generate PC Overview Excel Sheet of all Clusters
def getsnapshotDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating Snapshot Details:")

    ws = wb.create_sheet("Snapshot Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:C{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'Snapshot Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+2, '7855FA', ws)
    r += 1

    for entity in clustersData['entities']:
        
        if entity['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = entity['status']['name']
        print(f"Generating Snapshot Details for {clus_name}")
        ws.merge_cells(f'A{r}:C{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} Snapshot Details'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+2, 'A9D0F5', ws)
        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = entity['metadata']['uuid']
    

        clus_ip = entity['status']['resources']['network']['external_ip']

        ws.append(["Snapshot Name","Creation Date and Time", "VM Name"])
        styles.applyFont(r, c, c+3, ws)
        r += 1
        
        url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v2.0/snapshots"

        snapshots = apiCall.getApiCall(url, peAuthKey, clus_uuid)
        
        if snapshots == 0 or snapshots == {} or snapshots == 'null':
          continue
                        
        for s in snapshots['entities']:

            sname = s['snapshot_name']
            
            smicros = s['created_time']
            seconds = smicros / 1_000_000
            sdate = datetime(1970, 1, 1) + timedelta(seconds=seconds)
            vm_uuid = s['vm_uuid']

            url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v2.0/vms/{vm_uuid}"
            vm = apiCall.getApiCall(url, peAuthKey, clus_uuid)

            if vm == "-":
              vm_name = '-'

            else:
              vm_name = vm['name']


            ws.append([sname, sdate, vm_name])
            styles.applyFont(r, c, c+2, ws)
            r += 1
        
    styles.align(ws)
    styles.border(1, r-1, 1, 3, ws)