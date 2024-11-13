import styles
from openpyxl.styles import Font


#Generate PC Overview Excel Sheet of all Clusters 
def getLicenseDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    ws = wb.create_sheet("License Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:N{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'License Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+13, '7855FA', ws)
    r += 1

    

    ws.append(['License Id', 'Tier', 'License Class', 'Tag', 'Model Number', 'Licensing Metric', 'Purchased', 'Used / Purchased', 'PO Number', 'Days Until Start', 'Start Date', 'Expiration Date', 'SW Asset Serial Number', 'Associated Clusters'])
    styles.applyFont(r, c, c+13, ws)
    r += 1
                    
    styles.align(ws)
    styles.border(1, r-1, 1, 14, ws)