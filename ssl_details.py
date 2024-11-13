import apiCall, styles
from openpyxl.styles import Font


#Generate PC Overview Excel Sheet of all Clusters
def getsslDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating SSL Details:")

    ws = wb.create_sheet("SSL Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:C{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'SSL Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+2, '7855FA', ws)
    r += 1

    for entity in clustersData['entities']:
        
        if entity['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = entity['status']['name']
        print(f"Generating SSL Details for {clus_name}")
        ws.merge_cells(f'A{r}:C{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} SSL Details'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+2, 'A9D0F5', ws)
        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = entity['metadata']['uuid']
    

        clus_ip = entity['status']['resources']['network']['external_ip']

        ws.append(["Organization Name","Key Type", "Expire Date"])
        styles.applyFont(r, c, c+2, ws)
        r += 1
        
        url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v1/keys/pem"

        
        ssl = apiCall.getApiCall(url, peAuthKey, clus_uuid)
        
                        
        orgName = ssl['organizationName']
        keyType = ssl['keyType']
        exp = ssl['expiryDate']


        ws.append([orgName, keyType, exp])
        styles.applyFont(r, c, c+2, ws)
        r += 1
        
    styles.align(ws)
    styles.border(1, r-1, 1, 3, ws)