import apiCall, styles, hosts
from openpyxl.styles import Font


#Generate PC Overview Excel Sheet of all Clusters 
def getHostDiskDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating Host Disks Details:")

    ws = wb.create_sheet("Host Disks")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:J{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'Host Disk Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+9, '7855FA', ws)
    r += 1

    for clus in clustersData['entities']:
        
        if clus['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = clus['status']['name']
        print(f"Generating Host Disk Details for {clus_name}")
        ws.merge_cells(f'A{r}:J{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} Host Disks'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+9, 'A9D0F5', ws)
        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = clus['metadata']['uuid']
    

        clus_ip = clus['status']['resources']['network']['external_ip']
       

        ws.append(["Disk ID","Serial Number", "Host Name","Hypervisor IP", "Tier","Status","Disk Usage", "Disk IOPS", "Disk IO B/W", "Disk Avg IO Latency"])
        styles.fillColor(r, r, c, c+9, 'D3D3D3', ws)
        styles.applyFont(r, c, c+9, ws)
        r += 1

        url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v1/disks?&sortCriteria=id&projection=stats%2Calerts"
      
        hostDisks = apiCall.getApiCall(url, peAuthKey, clus_uuid)

        #Loop through each host

        for entity in hostDisks['entities']:
                                    
                    index = entity["id"].find('::')
                    
                    id = entity["id"][index + 2:]

                    serialNo = entity['diskHardwareConfig']['serialNumber']

                    hostName = hosts.hostsIPs[entity['hostName']]

                    hostIP = entity['hostName']

                    tier = entity['storageTierName']

                    if entity['online']:
                       status = 'Online'
                    
                    else:
                       status = 'Offline'

                    disk_usage = round(int(entity['usageStats']['storage.usage_bytes']) / 1073741824, 2)
                    disk_tot = round(int(entity['usageStats']['storage.usage_bytes']) / 1099511627776, 2)

                   
                    disk_iops = int(entity["stats"]["num_iops"])

                    disk_bw = int(entity["stats"]["io_bandwidth_kBps"])

                    disk_latency = round(int(entity["stats"]["avg_io_latency_usecs"]) / 1000, 2)


                    ws.append([id, serialNo, hostName, hostIP, tier, status, f"{disk_usage} GiB of {disk_tot} TiB ", disk_iops, disk_bw, disk_latency])
                    styles.applyFont(r, c, c+9, ws)
                    r += 1
                    
    styles.align(ws)
    styles.border(1, r-1, 1, 10, ws)