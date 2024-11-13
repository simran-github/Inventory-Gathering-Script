import apiCall, styles
from openpyxl.styles import Font
from datetime import datetime,timedelta

def microseconds_to_date(microseconds):
    seconds, microseconds = divmod(microseconds, 1000000)
    minutes, seconds = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)
    days, hours = divmod(hours, 24)
    return datetime(1970, 1, 1) + timedelta(days=days, hours=hours, minutes=minutes, seconds=seconds, microseconds=microseconds)


def bytes_to_gib(bytes_value):
    gib = bytes_value / (1024 ** 3)
    return gib

def getReplicationDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating Replication Details:")

    ws = wb.create_sheet("Replication Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:J{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'Replication Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+9, '7855FA', ws)
    r += 1
    
    for entity in clustersData['entities']:
        
        if entity['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = entity['status']['name']
        print(f"Generating Replication Details for {clus_name}")
        ws.merge_cells(f'A{r}:J{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} Protection Domains'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+9, 'A9D0F5', ws)
        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = entity['metadata']['uuid']
    

        clus_ip = entity['status']['resources']['network']['external_ip']

        url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v2.0/protection_domains"
        pdData = apiCall.getApiCall(url, peAuthKey, clus_uuid)

        ws.append(['Mode', 'Name', 'Remote Sites', 'Entity Count','Next Snapshot Time','Snapshot Exclusive Usage','B/w Used (Tx) (Kbps)','B/W Used(Rx) (Kbps)','Ongoing Replication Count','Pending Replication Count'])
        styles.fillColor(r, r, c, c+9, 'D3D3D3', ws)
        styles.applyFont(r, c, c+9, ws)

        r = r+1

        for pdData in pdData['entities']:

            pd_mode = pdData['active']

            pd_name = pdData['name']

            pd_remote_sites = ""

            if pdData['remote_site_names']:     
                for x in pdData['remote_site_names']:
                
                  pd_remote_sites = pd_remote_sites + x + ","


            pd_entity_count = len(pdData['vms'])

            if pdData['next_snapshot_time_usecs']:
                pd_next_snapshot_time = microseconds_to_date(pdData['next_snapshot_time_usecs'])
            else:
                pd_next_snapshot_time = ''

            
            if pdData['usage_stats']['dr.exclusive_snapshot_usage_bytes']:
                input_bytes = int(pdData['usage_stats']['dr.exclusive_snapshot_usage_bytes'])
            
                pd_exclusive_usage = bytes_to_gib(input_bytes)
            else:
                pd_exclusive_usage = ''

            
            pd_transmitted = pdData['stats']['replication_transmitted_bandwidth_kBps']

            pd_received = pdData['stats']['replication_received_bandwidth_kBps']

            pdData['stats']['replication_transmitted_bandwidth_kBps']
        
            pd_ongoing_rep_count = pdData['ongoing_replication_count']

            pd_pending_rep_count = pdData['pending_replication_count']

            

            ws.append([pd_mode, pd_name,pd_remote_sites,pd_entity_count,pd_next_snapshot_time,pd_exclusive_usage,pd_transmitted,pd_received,pd_ongoing_rep_count,pd_pending_rep_count])

            styles.applyFont(r, c, c+9, ws)
            r = r + 1

    styles.align(ws)
    styles.border(1, r-1, 1, 10, ws)