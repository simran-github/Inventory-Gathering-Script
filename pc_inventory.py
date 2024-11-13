import styles, apiCall
from openpyxl.styles import Font


def getPCInventory(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating Inventory Details:")

    ws = wb.active
    ws.title = "PC Inventory"
    rp = 1
    cp = 1
    

    ws.merge_cells('A1:P1')  
    cell = ws.cell(row=rp, column=1)  
    cell.value = f'PC Clusters Overview ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(rp, rp, cp, cp+15, '7855FA', ws)
    rp += 1


    styles.fillColor(rp, rp, cp, cp+15, '131313', ws)
    ws.merge_cells('A2:K2')
    rp += 1  


    ws.append(["Owner Team", "Cluster Name", "Cluster Virtual IP", "Cluster Data Services IP", "License Type", "Cluster UUID", "Blocks", "Nodes", "AOS Version", "Hypervisor Type", "Hypervisor Version", "NCC Version", "Foundation Version", "LCM Version", "VM Count", "Pulse Status"])
    ws.row_dimensions[rp].height = 30
    styles.fillColor(rp, rp, cp, cp+15, 'D3D3D3', ws)
    styles.border(rp, rp, cp, cp+15, ws)
    rp += 1
    
    


    for entity in clustersData['entities']:

        #Eliminate Prism central entry from the clusters list

        if entity['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
            continue

        #Get Cluster IPs
        clus_ip = "-"
        if 'external_ip' in entity['status']['resources']['network']:
            clus_ip = entity['status']['resources']['network']['external_ip']

        clus_ds_ip = "-"
        if 'external_data_services_ip' in entity['status']['resources']['network']:
            clus_ds_ip = entity['status']['resources']['network']['external_data_services_ip']

        clus_name = entity['status']['name']
        print(f"Generating Inventory Details for {clus_name}")
        if passw == '2' or passw == '1':
            peAuthKey = authKeys

        if passw == '3':
            peAuthKey = authKeys[clus_name]


        #Get Cluster UUID
        clus_uuid = entity['metadata']['uuid']

        #Get Cluster License
        url = f"https://{pc_name}:9440/PrismGateway/services/rest/v1/license?showAllClusters=true"
        licen_res = apiCall.getApiCall(url, pcAuthKey, clus_uuid)

        license_type = ""

        for l in licen_res['licenseInfoDTO']['pcDetailsDTO']['clusters']:

            if l['name'] == clus_name:

                license_type = l['licenseDetails'][0]['name']
                


        #Get No. of Blocks in the cluster
        url = f"https://{pc_name}:9440/PrismGateway/services/rest/v1/clusters?&proxyClusterUuid={clus_uuid}"
        block_res = apiCall.getApiCall(url, pcAuthKey, clus_uuid)


        for i in block_res['entities']:
        
            block_no = len(i['blockSerials'])


        #Get No. of Nodes in the cluster
        nodes = entity['status']['resources']['nodes']['hypervisor_server_list']

        ahv = nodes[0]['type']

        if ahv == 'VMWARE':
            node_count = len(nodes)

        else:
            node_count = len(nodes) - 1

        #Get AOS Version
        aos = entity['status']['resources']['config']['software_map']['NOS']['version']
        
        #Get NCC Version
        ncc = entity['status']['resources']['config']['software_map']['NCC']['version']

        ncc = ncc[len("ncc-"):]
        
        #Get Hypervisor type 
        ahv = nodes[0]['type']

        if nodes[0]['ip'] == '127.0.0.1':
            ahv_version = nodes[1]['version']
        else:
            ahv_version = nodes[0]['version']


        #Get Cluster Virtual IP

        url = f"https://{clus_ip}:9440/api/lcm/v4.0.a1/resources/entities"

        lcmres = apiCall.apiV4(url, peAuthKey)
        foundation = ""
        lcm = ""
      
        for l in lcmres['entities']:

            if l['entityModel'] == "Foundation":
          
                foundation = l["version"]

        

        url =  f"https://{clus_ip}:9440/api/lcm/v4.0.a1/resources/config"

        lcmVer = apiCall.apiV4Conf(url, peAuthKey)

        lcm = lcmVer['data']['lcmVersion']

        #Get VM Count in the Cluster
        vm_count = 0

        url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v2.0/vms"
        
        vm_res = apiCall.getApiCall(url, peAuthKey, clus_uuid)
        vm_count = vm_res['metadata']['grand_total_entities']

        url = f"https://{pc_name}:9440/PrismGateway/services/rest/v1/pulse?proxyClusterUuid={clus_uuid}"
        pulse_res = apiCall.getApiCall(url, pcAuthKey, clus_uuid)

        if pulse_res["enable"]:
            pulse = "YES"
              
        else:
            pulse = "NO"
        
        ws.append(["", clus_name, clus_ip, clus_ds_ip, license_type, clus_uuid, block_no , node_count, aos, ahv, ahv_version, ncc, foundation, lcm, vm_count, pulse])

        styles.applyFont(rp, cp, cp+15, ws)
        rp += 1

    styles.align(ws)
    styles.border(1, rp-1, 1, 16, ws)