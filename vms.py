import apiCall, styles
from openpyxl.styles import Font


#Generate PC Overview Excel Sheet of all Clusters 

def getVMsDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating VM Details")

    ws = wb.create_sheet("VM Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:M{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'VM Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+12, '7855FA', ws)
    r += 1

    for entity in clustersData['entities']:
        
        if entity['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = entity['status']['name']
        print(f"Generating VM Details for {clus_name}")
        ws.merge_cells(f'A{r}:M{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} VMs'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+12, 'A9D0F5', ws)
        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = entity['metadata']['uuid']
    

        clus_ip = entity['status']['resources']['network']['external_ip']
       



        ws.append(["VM Name", "Host", "IP Address", "Cores", "Memory Capacity", "Storage", "CPU Usage", "Memory Usage", "Controller Read IOPS", "Controller Write IOPS", "Controller IO Bandwidth", "Controller IO Latency", "Flash Mode"])
        styles.fillColor(r, r, c, c+12, 'D3D3D3', ws)
        styles.applyFont(r, c, c+12, ws)
        r += 1

        url = f"https://{clus_ip}:9440/PrismGateway/services/rest/v1/vms?&sortCriteria=vm_name&searchAttributeList=vm_uuid&projection=stats%2CbasicInfo%2Calerts&filterCriteria=is_control_domain!%3D1%3Bis_cvm%3D%3D0"
      
        vmsData = apiCall.getApiCall(url, peAuthKey, clus_uuid)

        #Loop through each host

        for v in vmsData['entities']:
                                

                    vmName = v['vmName']
                    host = v["hostName"]
                    ipAddr = ""
                    for ip in v['ipAddresses']:
                      ipAddr = ipAddr + ip

                    cores = v['numVCpus']

                    mem = round(int(v['memoryCapacityInBytes']) / 1073741824, 2)

                    if v['diskCapacityInBytes'] == None:
                        storage = '-'
                    else:
                        storage = round(int(v['diskCapacityInBytes']) / 1073741824, 2)
                        


                    cpu_usage = round(int(v["stats"]["hypervisor_cpu_usage_ppm"]) / 10000, 2)

                    mem_usage = round(int(v["stats"]["memory_usage_ppm"]) / 10000, 2)

                    storage_usage = round(int(v["stats"]["controller_user_bytes"]) / 1073741824, 2)

                    contr_read_iops = int(v["stats"]["controller_num_read_iops"])

                    contr_write_iops = int(v["stats"]["controller_num_write_iops"])

                    control_io_bw = int(v["stats"]["controller_io_bandwidth_kBps"])

                    contr_io_latency = round(int(v["stats"]["controller_avg_io_latency_usecs"]) / 1000, 2)

                    if v['vmFeatures']['FLASH_MODE']:
                        flash = 'ON'
                    else:
                        flash = 'OFF'


                    ws.append([vmName, host, ipAddr, cores, f"{mem} GiB", f"{storage_usage} GiB / {storage} GiB", f"{cpu_usage} %", f"{mem_usage} %", contr_read_iops, contr_write_iops, f"{control_io_bw} KBps", f"{contr_io_latency} ms", flash])
                    styles.applyFont(r, c, c+12, ws)
                    r += 1
                    
    styles.align(ws)
    styles.border(1, r-1, 1, 13, ws)