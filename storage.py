import apiCall, styles
from openpyxl.styles import Font

def getStorageDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating Storage Details")

    ws = wb.create_sheet("Storage Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:L{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'Storage Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+11, '7855FA', ws)
    r += 1



    for entity in clustersData['entities']:
        
        if entity['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = entity['status']['name']
        print(f"Generating Storage Details for {clus_name}")
        ws.merge_cells(f'A{r}:L{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} Storage Details'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+11, 'A9D0F5', ws)
        r += 1

        ws.merge_cells(f'A{r}:L{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} Storage Containers'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+11, 'F0F3FF', ws)
        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = entity['metadata']['uuid']
    

        clus_ip = entity['status']['resources']['network']['external_ip']
       

        url = f'https://{clus_ip}:9440/PrismGateway/services/rest/v2.0/storage_containers'
        storageData = apiCall.getApiCall(url,peAuthKey, clus_uuid)
    

        ws.append(["Name","Replication factor", "Compression","Deduplication","Erasure Coding","Free(Physical TiB)", "Used(TiB)", "Reserved (TiB)", "Max Capacity (TiB)", "Controller IOPs", "Controller IO Bandwidth(Kbps)", "Controller IO Latency(ms)"])
        styles.fillColor(r, r, c, c+11, 'D3D3D3', ws)
        styles.applyFont(r, c, c+11, ws)
        r += 1

        for containr in storageData['entities']:
              
                container_name = containr['name']

                container_rep_factor = containr["replication_factor"]

                container_compression = containr['compression_enabled']

                container_dedup = containr['on_disk_dedup']

                container_erasure_coding = containr['erasure_code']

                container_free_logical = round(int(containr['usage_stats']['storage.capacity_bytes'])/1099511627776, 2) 

                container_used = round(int(containr['usage_stats']['storage.user_disk_physical_usage_bytes'])/1073741824,2)

                container_reserved = round(int(containr['usage_stats']['storage.container_reserved_capacity_bytes']))

                container_max = round(int(containr['max_capacity'])/1099511627776, 2)

                container_controller_iops = containr['stats']['controller_num_iops']

                container_io_bw = containr['stats']['controller_io_bandwidth_kBps']

                if containr["stats"]["controller_avg_io_latency_usecs"] == -1:
                    container_io_latency = "-"
                else:
              
                 container_io_latency = round(int(containr["stats"]["controller_avg_io_latency_usecs"]) / 1000, 2)


                ws.append([container_name,container_rep_factor ,container_compression, container_dedup,container_erasure_coding, container_free_logical, container_used,container_reserved,container_max, container_controller_iops,container_io_bw,container_io_latency])
                styles.applyFont(r, c, c+13, ws)
                r += 1

        ws.merge_cells(f'A{r}:I{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} Storage Pools'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+8, 'F0F3FF', ws)
        r += 1

        url = f'https://{clus_ip}:9440/PrismGateway/services/rest/v1/storage_pools?&sortCriteria=storage_pool_name&_=1717960272539&projection=stats%2Calerts'
        storagePoolsData = apiCall.getApiCall(url,peAuthKey, clus_uuid)

        ws.append(["Name","Disks", "Free(Physical)","Used(Physical) GiB","Max Capacity(Physical)","Disk IOPS", "Disk IO B/W", "Disk Avg IO Latency"])
        styles.fillColor(r, r, c, c+8, 'D3D3D3', ws)
        styles.applyFont(r, c, c+8, ws)
        r += 1

        for stor_pool in storagePoolsData['entities']:

            pool_name = stor_pool['name']

            pool_disks = len(stor_pool['disks'])

            pool_free = round(int(stor_pool['usageStats']['storage.free_bytes'])/1099511627776,2)

            pool_used = round(int(stor_pool['usageStats']['storage.usage_bytes'])/1073741824,2)

            pool_max_capacity = round(int(stor_pool['usageStats']['storage.capacity_bytes'])/1099511627776,2)

            pool_iops = int(stor_pool["stats"]["num_iops"])

            pool_disk_bw = int(stor_pool["stats"]["io_bandwidth_kBps"])

            pool_disk_latency = round(int(stor_pool["stats"]["avg_io_latency_usecs"]) / 1000, 2)


            ws.append([pool_name,pool_disks,pool_free,pool_used,pool_max_capacity,pool_iops,pool_disk_bw,pool_disk_latency])
            styles.applyFont(r, c, c+8, ws)
            r += 1
    
    styles.align(ws)
    styles.border(1, r-1, 1, 12, ws)