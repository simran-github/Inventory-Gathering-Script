import apiCall, styles
from openpyxl.styles import Font


#Generate PC Overview Excel Sheet of all Clusters 
def getvlanDetails(pc_name,clustersData, pcAuthKey, authKeys, passw, wb):

    print("Generating Network Details:")

    ws = wb.create_sheet("VLAN and Switch Details")
    r = 1
    c = 1

    ws.merge_cells(f'A{r}:F{r}')
    cell = ws.cell(row=r, column=1)
    cell.value = f'VLAN Details ({pc_name})'
    cell.font = Font(name= 'Helvetica', size=16)
    ws.row_dimensions[1].height = 50
    styles.fillColor(r, r, c, c+5, '7855FA', ws)
    r += 1

    for entity in clustersData['entities']:
        
        if entity['status']['resources']['config']['service_list'][0] == "PRISM_CENTRAL":
          continue

        clus_name = entity['status']['name']
        print(f"Generating Network Details for {clus_name}")
        ws.merge_cells(f'A{r}:F{r}')
        cell = ws.cell(row=r, column=1)
        cell.value = f'{clus_name} VLANS'
        cell.font = Font(name= 'Helvetica', size=16)
        ws.row_dimensions[r].height = 30
        styles.fillColor(r, r, c, c+5, 'A9D0F5', ws)
        

        cell = ws.cell(row=r, column=8)
        cell.value = f'Name'

        cell = ws.cell(row=r, column=9)
        cell.value = f'Bridge'

        cell = ws.cell(row=r, column=10)
        cell.value = f'MTU (bytes)'

        cell = ws.cell(row=r, column=11)
        cell.value = f'Bond Type'

        styles.fillColor(r, r, 8, 11, 'D3D3D3', ws)

        r += 1

        if passw == '2' or passw == '1':
          peAuthKey = authKeys

        if passw == '3':
          peAuthKey = authKeys[clus_name]
      

        clus_uuid = entity['metadata']['uuid']
    

        clus_ip = entity['status']['resources']['network']['external_ip']

        if entity['status']['resources']['nodes']['hypervisor_server_list'][0]['type'] != 'VMWARE':

            url = f"https://{pc_name}:9440/api/networking/v2.a1/dvs/virtual-switches?&proxyClusterUuid={clus_uuid}"
      
            vswitchList = apiCall.getApiCall(url, pcAuthKey, clus_uuid)
       
            for v in vswitchList:

                vsName = v['data']['name']
                bridge = v['data']['clusters'][0]['hosts'][0]['internalBridgeName']
                mtu = v['data']['mtu']
                bond = v['data']['bondMode']

                cell = ws.cell(row=r, column=8)
                cell.value = f'{vsName}'

                cell = ws.cell(row=r, column=9)
                cell.value = f'{bridge}'

                cell = ws.cell(row=r, column=10)
                cell.value = f'{mtu}'

                cell = ws.cell(row=r, column=11)
                cell.value = f'{bond}'
                styles.border(r-1, r, 8, 11, ws)
            
                styles.fillColor(r, r, c, c+5, 'D3D3D3', ws)
                styles.applyFont(r, c, c+5, ws)
                r += 1
            
        ws.append(["Network Name","Virtual Switch", "VLAN ID","Used IP Addresses", "Free IPs in Subnets","Free IPs in Pool"])
        styles.applyFont(r, c, c+5, ws)
        r += 1
        
        if entity['status']['resources']['nodes']['hypervisor_server_list'][0]['type'] != 'VMWARE':
            url = f"https://{clus_ip}:9440/api/nutanix/v0.8/networks"
        
            vlanList = apiCall.getApiCall(url, peAuthKey, clus_uuid)

            #Loop through each host

            for entity in vlanList['entities']:
                
                vlan = '-'
                if 'name' in entity:
                    vlan = entity["name"]

                vlanID = '-' 
                if 'vlanId' in entity:
                    vlanID = entity["vlanId"]
                
                vs = '-'
                if 'vswitchName' in entity:
                    vs = entity['vswitchName']

                if entity['ipConfig']['assignedIps'] == -1: 

                    useIPs = 'N/A'

                else:
                    useIPs = entity['ipConfig']['assignedIps']

                if entity['ipConfig']['freeIps'] == -1:
                    freeIps = 'N/A'
                
                else:
                    freeIps = entity['ipConfig']['freeIps']

                if entity['ipConfig']['pool'] == []:
                    poolFree = 'N/A'
                
                else:
                    poolFree = 0
                    for p in entity['ipConfig']['pool']:
                        poolFree = poolFree + p['numFreeIps']

                ws.append([vlan, vs, vlanID, useIPs, freeIps, poolFree])
                styles.applyFont(r, c, c+5, ws)
                r += 1
        else:
            ws.append(["Cluster is ESXi. Cannot fetch network details"])    
            r += 1        
    styles.align(ws)
    styles.border(1, r-1, 1, 6, ws)